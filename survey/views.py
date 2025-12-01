import logging 
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.core.paginator import Paginator
from django.views.generic import *
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required,user_passes_test
from django.utils import timezone 
from django.utils.timesince import timesince
from django.contrib import messages
from django.db.models import Count, Q , Avg
from django.db import transaction
from django.contrib.auth.models import User, Group 
from django.contrib.sessions.models import Session
from django.http import HttpResponse
from django.conf import settings
from datetime import timedelta , datetime
from openpyxl.utils import get_column_letter
from .forms import *
from .models import * 
from .utils import *
import openpyxl
import csv
import json

# สร้าง Logger Instance
logger = logging.getLogger(__name__)

@login_required
def after_login_view(request):
    user = request.user  
    if user.is_superuser:
        # ถ้าเป็น Admin -> ไป Dashboard ใหญ่
        return redirect('survey:dashboard')
    else:
        # ถ้าเป็น Manager (คนทั่วไป) -> ไป Dashboard ส่วนตัว
        return redirect('manager:dashboard')
# --- Auxiliary Functions ---

def custom_logout_view(request):
    # 1. ล้างข้อมูลเซสชันอื่นๆ ที่กำหนดเอง (ถ้ามี)
    if 'patient_info' in request.session:
        del request.session['patient_info']
    
    # 2. ออกจากระบบ (ล้าง User Session และ Cookies)
    logout(request) 
    
    # 3. เปลี่ยนเส้นทางไปยังหน้าหลัก
    return redirect('homepage')

def is_superuser(user):
    return user.is_superuser

def get_summary_context():
    return {
        'total_service_points': ServicePoint.objects.count(),
        'total_service_groups': ServiceGroup.objects.count(),
    }

def _get_point_map():
    """สร้างแผนที่ (JSON) ของ {Group: [Points]} สำหรับ Dropdown 2 ชั้น"""
    point_map = {}
    groups = ServiceGroup.objects.prefetch_related('service_points')
    
    for group in groups:
        point_map[group.id] = [
            {'id': point.id, 'name': point.name}
            for point in group.service_points.all().order_by('name') 
        ]
    return point_map

# --- Mixins ---
class SuperuserRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

# --- General Views ---
@login_required
def index(request):
    return render(request, "index.html")

def Home(request) :
    return render(request, 'survey/home.html')

# --- Dashboard View --
@login_required
def dashboard_view(request):
    user = request.user
    base_service_points = ServicePoint.objects.all()
    managers_list = User.objects.none() 
    base_service_points = ServicePoint.objects.none() # กำหนดค่าเริ่มต้นกันพลาด

    if user.is_authenticated:
        if not user.is_superuser:
            # กรณี User ธรรมดา: ดูเฉพาะจุดที่ตัวเองดูแล
            base_service_points = user.managed_points.all()
            managers_list = User.objects.filter(id=user.id).prefetch_related('managed_points')
        else:
            # --- แก้ไขตรงนี้ (Superuser) ---
            base_service_points = ServicePoint.objects.all()
            
            # วิธีเดิม: พึ่งพา Group 'Managers' (ถ้าไม่มี Group ข้อมูลจะไม่ขึ้น)
            # วิธีใหม่: ดึง User ทุกคนที่มีความสัมพันธ์กับ managed_points (มีจุดบริการที่ดูแลอยู่อย่างน้อย 1 จุด)
            managers_list = User.objects.filter(
                managed_points__isnull=False
            ).distinct().prefetch_related('managed_points')

    # A2: กรองตามวันที่ (Date Filter)
    today = timezone.now().date()
    start_date_default = today - timedelta(days=today.weekday()) # จันทร์
    end_date_default = start_date_default + timedelta(days=6) # อาทิตย์
    
    end_date_str = request.GET.get('end_date', end_date_default.strftime('%Y-%m-%d'))
    start_date_str = request.GET.get('start_date', start_date_default.strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = start_date_default
        end_date = end_date_default
        
    end_date_for_query = end_date + timedelta(days=1)
    
    # --- B. สร้าง Queryset หลักที่ "กรองแล้ว" ---
    filtered_responses_for_charts_and_ticker = Response.objects.filter(
        service_point__in=base_service_points, 
        submitted_at__gte=start_date, 
        submitted_at__lt=end_date_for_query 
    )

    # --- C. คำนวณข้อมูล ---
    total_active_questions = Question.objects.filter(
        survey__status=Survey.Status.ACTIVE
    ).count()
    total_responses = filtered_responses_for_charts_and_ticker.count()

    # กรองโดยใช้ 'filtered_responses_for_charts_and_ticker'
    all_service_points_with_counts = base_service_points.annotate(
        response_count=Count('response', filter=Q(response__in=filtered_responses_for_charts_and_ticker))
    ).order_by('-response_count')

    # Card 3: กราฟแท่งรายสัปดาห์ (ซ้ายกลาง)
    date_labels = []
    day_counts_dict = {}
    current_date = start_date
    while current_date <= end_date:
        date_labels.append(current_date.strftime('%a')) 
        day_counts_dict[current_date] = 0
        current_date += timedelta(days=1)

    response_times = filtered_responses_for_charts_and_ticker.values_list('submitted_at', flat=True)

    for submitted_at_utc in response_times:
        local_time = timezone.localtime(submitted_at_utc)
        date_only = local_time.date() 
        if date_only in day_counts_dict:
            day_counts_dict[date_only] += 1
    bar_data_weekly = [day_counts_dict[day] for day in sorted(day_counts_dict.keys())]

    pie_data_all = base_service_points.annotate(
        response_count=Count('response', filter=Q(response__in=filtered_responses_for_charts_and_ticker))
    ).filter(response_count__gt=0).order_by('-response_count')
    pie_labels = [sp.name for sp in pie_data_all]
    pie_data = [sp.response_count for sp in pie_data_all]

    recent_feedback = ResponseAnswer.objects.filter(
        response__in=filtered_responses_for_charts_and_ticker,
        answer_text__isnull=False  # ต้องมีค่าในคอลัมน์นี้
    ).exclude(
        answer_text__exact=''      # ต้องไม่ใช่ข้อความว่างเปล่า
    ).select_related(
        'response__service_point'
    ).order_by('-response__submitted_at')[:5]
    
    # --- D. ส่งข้อมูลทั้งหมดไปที่ Template ---
    context = {
        'total_responses': total_responses,
        'total_service_points_in_view': base_service_points.count(),
        'total_active_questions': total_active_questions,
        'all_service_points_with_counts': all_service_points_with_counts,
        'bar_labels_weekly': json.dumps(date_labels),
        'bar_data_weekly': json.dumps(bar_data_weekly),
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
        'managers_list': managers_list,
        'recent_feedback': recent_feedback,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }

    return render(request, 'survey/dashboard.html', context)

# ========== 1. Service Point Views (จุดบริการ) ==========
@login_required
@user_passes_test(is_superuser)
def service_point_list_view(request):
    queryset = ServicePoint.objects.select_related('group').prefetch_related('managers').order_by('code')
    search_query = request.GET.get('q', '')
    group_id = request.GET.get('group_id', '')

    if search_query:
        queryset = queryset.filter(Q(name__icontains=search_query) | Q(code__icontains=search_query))
    
    if group_id:
        queryset = queryset.filter(group_id=group_id)

    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    groups = ServiceGroup.objects.annotate(point_count=Count('service_points')).order_by('name')

    context = {
        'title': 'จัดการจุดบริการและกลุ่มภารกิจ',
        'page_obj': page_obj,
        'all_groups': ServiceGroup.objects.all().order_by('name'),
        'search_query': search_query,
        'group_id': group_id,
        'groups': groups, 
    }
    context.update(get_summary_context())
    return render(request, 'survey/service_point_list.html', context)

@login_required
@user_passes_test(is_superuser)
def service_point_create_view(request):
    if request.method == 'POST':
        form = ServicePointForm(request.POST)
        if form.is_valid():
            point = form.save()
            messages.success(request, f'เพิ่มจุดบริการ "{point.name}" สำเร็จ')
            return redirect('survey:service_point_list')
    else:
        form = ServicePointForm()

    context = {
        'title': 'เพิ่มจุดบริการใหม่',
        'form': form,
    }
    context.update(get_summary_context())
    return render(request, 'survey/service_point_form.html', context)

@login_required
@user_passes_test(is_superuser)
def service_point_edit_view(request, pk):
    point = get_object_or_404(ServicePoint, pk=pk)
    if request.method == 'POST':
        form = ServicePointForm(request.POST, instance=point)
        if form.is_valid():
            form.save()
            messages.success(request, f'อัปเดตจุดบริการ "{point.name}" สำเร็จ')
            return redirect('survey:service_point_list')
    else:
        form = ServicePointForm(instance=point)

    context = {
        'title': f'แก้ไขจุดบริการ: {point.name}',
        'form': form,
        'point': point,
    }
    context.update(get_summary_context())
    return render(request, 'survey/service_point_form.html', context)

@login_required
@user_passes_test(is_superuser)
def service_point_delete_view(request, pk):
    point = get_object_or_404(ServicePoint, pk=pk)
    try:
        point.delete()
        messages.success(request, f'ลบจุดบริการ "{point.name}" สำเร็จ')
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาดในการลบ: {e}')
    return redirect('survey:service_point_list')

# ========== 2. Service Group Views (กลุ่มภารกิจ) ==========
@login_required
@user_passes_test(is_superuser)
def service_group_create_view(request):
    if request.method == 'POST':
        form = ServiceGroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            messages.success(request, f'เพิ่มกลุ่มภารกิจ "{group.name}" สำเร็จ')
            return redirect('survey:service_point_list') 
    else:
        form = ServiceGroupForm()

    context = {
        'title': 'เพิ่มกลุ่มภารกิจใหม่',
        'form': form,
    }
    context.update(get_summary_context())
    return render(request, 'survey/service_group_form.html', context)

@login_required
@user_passes_test(is_superuser)
def service_group_edit_view(request, pk):
    group = get_object_or_404(ServiceGroup, pk=pk)
    if request.method == 'POST':
        form = ServiceGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, f'อัปเดตกลุ่มภารกิจ "{group.name}" สำเร็จ')
            return redirect('survey:service_point_list') 
    else:
        form = ServiceGroupForm(instance=group)

    context = {
        'title': f'แก้ไขกลุ่มภารกิจ: {group.name}',
        'form': form,
        'group': group,
    }
    context.update(get_summary_context())
    return render(request, 'survey/service_group_form.html', context)

@login_required
@user_passes_test(is_superuser)
def service_group_delete_view(request, pk):
    group = get_object_or_404(ServiceGroup, pk=pk)
    if group.service_points.exists():
        messages.error(request, f'ไม่สามารถลบ "{group.name}" ได้ เพราะยังมีจุดบริการอยู่ในกลุ่มนี้')
        return redirect('survey:service_point_list')
        
    try:
        group.delete()
        messages.success(request, f'ลบกลุ่มภารกิจ "{group.name}" สำเร็จ')
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาดในการลบ: {e}')
    return redirect('survey:service_point_list')

##-----------ส่วนเพิ่มผู้ดูแล-----------------------
def get_manager_summary_context():
    manager_query = User.objects.filter(is_superuser=False)
    total_managers = manager_query.count()
    total_service_points = ServicePoint.objects.count()
    total_service_groups = ServiceGroup.objects.count() 
    all_groups = ServiceGroup.objects.all().order_by('name')
    manager_ids = set(manager_query.values_list('id', flat=True))
    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    online_manager_ids = []

    for session in sessions:
        data = session.get_decoded()
        user_id = data.get('_auth_user_id', None)
        if user_id and int(user_id) in manager_ids:
            online_manager_ids.append(int(user_id))
    
    online_managers = len(set(online_manager_ids))

    all_points = ServicePoint.objects.all().values('id', 'group_id')
    service_point_map = {
        sp['id']: sp['group_id'] if sp['group_id'] is not None else 'none' 
        for sp in all_points
    }

    return {
        'total_managers': total_managers,
        'total_service_points': total_service_points,
        'total_service_groups': total_service_groups,
        'online_managers': online_managers, 
        'service_groups': all_groups, 
        'service_point_group_map': json.dumps(service_point_map)
    }

@login_required 
@user_passes_test(is_superuser) 
def manager_list_view(request):
    # 1. รับค่า Filter
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '') # รับค่า status (online/offline)

    # 2. Query Managers เบื้องต้น
    managers_qs = User.objects.filter(is_superuser=False).prefetch_related('managed_points').order_by('username')

    # 3. กรอง Search Query
    if query:
        managers_qs = managers_qs.filter(
            Q(username__icontains=query) | Q(email__icontains=query)
        )

    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    online_user_ids = set()
    for session in sessions:
        data = session.get_decoded()
        uid = data.get('_auth_user_id')
        if uid:
            online_user_ids.add(int(uid))

    if status_filter == 'online':
        managers_qs = [m for m in managers_qs if m.id in online_user_ids]
    elif status_filter == 'offline':
        managers_qs = [m for m in managers_qs if m.id not in online_user_ids]

    context = {
        'managers': managers_qs, 
        'search_query': query,
        'status_filter': status_filter, # ส่งค่ากลับไปแสดงใน Dropdown
        'online_user_ids': online_user_ids,
    }
    context.update(get_manager_summary_context()) 
    
    return render(request, 'survey/manager_list.html', context)

@login_required
@user_passes_test(is_superuser)
def manager_create_view(request):
    if request.method == 'POST':
        form = ManagerCreateForm(request.POST) 
        if form.is_valid():
            user = form.save()
            messages.success(request, f'สร้างผู้ดูแล "{user.username}" สำเร็จ')
            return redirect('survey:manager_list')
    else:
        form = ManagerCreateForm() 
        
    context = {
        'form': form,
        'title': 'เพิ่มข้อมูลผู้ดูแล',
    }
    context.update(get_manager_summary_context()) 
    return render(request, 'survey/manager_form.html', context)

@login_required
@user_passes_test(is_superuser)
def manager_edit_view(request, pk):
    manager = get_object_or_404(User, pk=pk, is_superuser=False)
    
    if request.method == 'POST':
        form = ManagerEditForm(request.POST, instance=manager) 
        if form.is_valid():
            manager = form.save() 
            messages.success(request, f'อัปเดตข้อมูล "{manager.username}" สำเร็จ')
            return redirect('survey:manager_list')
    else:
        form = ManagerEditForm(instance=manager) 
        
    context = {
        'form': form,
        'manager': manager,
        'title': f'แก้ไขผู้ดูแล: {manager.username}',
    }
    context.update(get_manager_summary_context()) 
    return render(request, 'survey/manager_form.html', context)

@login_required
@user_passes_test(is_superuser)
def manager_delete_view(request, pk):
    manager = get_object_or_404(User, pk=pk, is_superuser=False)
    try:
        manager_name = manager.username
        manager.delete()
        messages.success(request, f'ลบผู้ดูแล "{manager_name}" สำเร็จ')
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาด: {e}')
    return redirect('survey:manager_list')


# --- CRUD: Survey ---
@login_required
@user_passes_test(is_superuser)
def survey_list_view(request):
    # 1. Queryset หลัก
    # ใช้ .select_related เพื่อลดการ Query ฐานข้อมูล (optimization)
    surveys = Survey.objects.annotate(
        question_count=Count('questions')
    ).select_related('service_point', 'service_point__group').order_by('-created_at')

    # --- 🔍 2. FILTER LOGIC ---
    search_query = request.GET.get('q')
    group_filter = request.GET.get('group')
    point_filter = request.GET.get('point')

    if search_query:
        surveys = surveys.filter(
            Q(title_th__icontains=search_query) | 
            Q(title_en__icontains=search_query)
        )
    if group_filter:
        surveys = surveys.filter(service_point__group_id=group_filter)
    if point_filter:
        surveys = surveys.filter(service_point_id=point_filter)

    # --- 📊 3. STATS (นับจากทั้งหมดในระบบ ไม่สน Filter) ---
    total_surveys_count = Survey.objects.count()
    active_surveys = Survey.objects.filter(status='ACTIVE').count()
    draft_surveys = Survey.objects.filter(status='DRAFT').count()
    total_questions = Question.objects.count()

    # --- 📄 4. PAGINATION ---
    paginator = Paginator(surveys, 5) # 👈 แสดงหน้าละ 5 รายการ
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # --- 🛠️ 5. DROPDOWN DATA & POINT MAP (สำหรับ JavaScript) ---
    groups = ServiceGroup.objects.all().order_by('name')
    service_points = ServicePoint.objects.all().order_by('name')
    
    # สร้าง Map สำหรับให้ JS จัดการ Dropdown จุดบริการอัตโนมัติ (แก้ไขปัญหา ID)
    point_map = {}
    for g in groups:
        # ใช้ related_name 'service_points' ที่ถูกกำหนดในโมเดล ServicePoint
        points = g.service_points.all().order_by('name') 
        point_map[g.id] = [{'id': p.id, 'name': p.name} for p in points]

    # --- 📝 6. FORM & MODAL LOGIC ---
    show_modal = False
    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by_user = request.user
            survey.save()
            messages.success(request, "สร้างแบบสอบถามเรียบร้อยแล้ว")
            return redirect('survey:survey_list') # 🚨 ปรับ URL name ให้ถูกต้อง
        else:
            show_modal = True
            messages.error(request, "เกิดข้อผิดพลาด กรุณาตรวจสอบข้อมูล")
    else:
        form = SurveyForm()

    context = {
        'page_title': 'จัดการแบบสอบถาม',
        'surveys': page_obj, 
        'form': form,
        'show_modal': show_modal,
        
        # Stats
        'total_surveys': total_surveys_count,
        'active_surveys': active_surveys,
        'draft_surveys': draft_surveys,
        'total_questions': total_questions,
        
        # Dropdowns
        'groups': groups,
        'service_points': service_points,
        # 📌 แปลง point_map เป็น JSON และส่งไปให้ Template
        'point_map_json': json.dumps(point_map), 
    }

    return render(request, 'survey/survey_list.html', context)

@login_required
@user_passes_test(is_superuser)
def survey_edit_view(request, pk):
    original_survey = get_object_or_404(Survey, pk=pk)

    if request.method == 'POST':
        # 🔴 สำคัญ: ส่ง instance เข้าไปเพื่อให้ Form รู้ว่ากำลังแก้ไขตัวไหน
        form = SurveyForm(request.POST, instance=original_survey)
        
        if form.is_valid():
            changed_data = form.changed_data
            new_status = form.cleaned_data.get('status')
            new_service_point = form.cleaned_data.get('service_point')
            
            # --- 🔴 STEP 1: CONSTRAINT CHECK (1 ACTIVE SURVEY PER SERVICE POINT) ---
            if new_status == 'ACTIVE':
                # ตรวจสอบว่ามีแบบสอบถามอื่น (ที่ไม่ใช่ตัวปัจจุบัน) ที่ Active อยู่แล้วหรือไม่
                if Survey.objects.filter(
                    service_point=new_service_point,
                    status='ACTIVE'
                ).exclude(pk=original_survey.pk).exists():
                    
                    messages.error(request, f"ไม่สามารถเปิดใช้งานได้: จุดบริการ **'{new_service_point.name}'** มีแบบสอบถามที่เปิดใช้งานอยู่แล้ว กรุณาเปลี่ยนสถานะของแบบสอบถามอื่นก่อน")
                    return redirect('survey:survey_list')

            # --- STEP 2: APPLY SAVE LOGIC ---
            
            # CASE A: Change only Status -> Update In-place (ไม่ต้องสร้างเวอร์ชันใหม่)
            if len(changed_data) == 1 and 'status' in changed_data:
                form.save()
                messages.success(request, "อัปเดตสถานะเรียบร้อยแล้ว")
            
            # CASE C: Change Content (หรืออื่นๆ) -> Create New Version
            else:
                try:
                    with transaction.atomic():
                        
                        # 2.1 [Optional Cleanup] ถ้าเวอร์ชันใหม่เป็น ACTIVE, ควรกำหนดให้เวอร์ชันเดิมเป็น DRAFT
                        #    (เพื่อทำความสะอาดประวัติ/ป้องกันความสับสน แม้ว่า Constraint จะจัดการแล้วก็ตาม)
                        if new_status == 'ACTIVE':
                             Survey.objects.filter(pk=original_survey.pk).update(status='DRAFT')

                        # 2.2 สร้าง Survey Object ใหม่
                        new_survey = form.save(commit=False)
                        new_survey.pk = None 
                        
                        # Version logic
                        current_ver = float(original_survey.version_number or 0)
                        new_survey.version_number = f"{int(current_ver) + 1}.0"
                        
                        new_survey.save()

                        # 2.3 Clone Questions
                        old_questions = original_survey.questions.all().order_by('order')
                        new_questions = [
                            Question(
                                survey=new_survey,
                                text_th=q.text_th,
                                text_en=q.text_en,
                                question_type=q.question_type,
                                order=q.order,
                                is_required=q.is_required
                            ) for q in old_questions
                        ]
                        Question.objects.bulk_create(new_questions)

                    messages.success(request, f"บันทึกเวอร์ชันใหม่ (v{new_survey.version_number}) เรียบร้อยแล้ว")

                except Exception as e:
                    messages.error(request, f"เกิดข้อผิดพลาด: {e}")
            
            return redirect('survey:survey_list')
        else:
            messages.error(request, "ข้อมูลไม่ถูกต้อง")
    
    return redirect('survey:survey_list')

class SurveyDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Survey
    success_url = reverse_lazy('survey:survey_list')

# --- Question Views ---
@login_required
@user_passes_test(lambda u: u.is_superuser)
def question_list_view(request, survey_id):
    survey = get_object_or_404(Survey, pk=survey_id)
    questions = survey.questions.all().order_by('order')
    
    # === ส่วนที่เพิ่ม: จัดการการสร้างคำถามในหน้านี้เลย ===
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            question.created_by_user = request.user
            question.save()
            
            messages.success(request, 'เพิ่มคำถามสำเร็จ')
            # Redirect กลับมาหน้าเดิม (เพื่อล้างค่า POST)
            return redirect('survey:question_list', survey_id=survey.id)
        else:
            # ถ้า Error ให้เปิด Modal ค้างไว้
            show_modal = True
    else:
        # เตรียมฟอร์มเปล่าสำหรับใส่ใน Modal
        last_order = survey.questions.all().count() + 1
        form = QuestionForm(initial={'order': last_order, 'survey': survey})
        show_modal = False
    # ==================================================

    return render(request, 'survey/question_list.html', {
        'survey': survey,
        'questions': questions,
        'form': form,           # ส่งฟอร์มไปที่ Template
        'show_modal': show_modal # ส่ง Flag เพื่อบอกว่าควรเปิด Modal ไหม (กรณี Error)
    })

class QuestionUpdateView(SuperuserRequiredMixin, UpdateView):
    model = Question
    form_class = QuestionForm
    
    def form_valid(self, form):
        messages.success(self.request, "แก้ไขคำถามเรียบร้อยแล้ว")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "เกิดข้อผิดพลาดในการแก้ไข กรุณาลองใหม่")
        return redirect('survey:question_list', survey_id=self.object.survey.id)

    def get_success_url(self):
        return reverse('survey:question_list', args=[self.object.survey.id])


class QuestionDeleteView(SuperuserRequiredMixin, DeleteView):
    model = Question
    template_name = 'survey/survey_confirm_delete.html' 
    def get_success_url(self):
        messages.success(self.request, "ลบคำถามเรียบร้อยแล้ว")
        return reverse('survey:question_list', args=[self.object.survey.id])

# --- Kiosk Views ---
def kiosk_welcome_view(request, service_point_id):
    service_point = get_object_or_404(ServicePoint, id=service_point_id)
    
    if 'patient_info' in request.session:
        del request.session['patient_info']
        
    # [แก้ไขคืนชีพ] กรองเอาเฉพาะ Active
    active_survey = Survey.objects.filter(
        service_point=service_point, 
        status=Survey.Status.ACTIVE
    ).order_by('-id').first()

    if not active_survey:
        return render(request, 'kiosk/kiosk_welcome.html', {
            'service_point': service_point,
            'active_survey': None
        })
    
    if request.method == 'POST':
        return redirect('survey:kiosk_user_info', service_point_id=service_point.id)

    context = {
        'service_point': service_point,
        'active_survey': active_survey
    }
    return render(request, 'kiosk/kiosk_welcome.html', context)
        
    if request.method == 'POST':
        # ถ้ามีการกดปุ่ม (POST) ให้พาไปหน้ากรอกข้อมูลผู้ป่วย
        return redirect('survey:kiosk_user_info', service_point_id=service_point.id)
    
    print(f"DEBUG: สรุปผล -> เลือกแบบสอบถาม ID: {active_survey.id}")
    context = {
        'service_point': service_point,
        'active_survey': active_survey
    }
    return render(request, 'kiosk/kiosk_welcome.html', context)

def kiosk_user_info_view(request, service_point_id):
    service_point = get_object_or_404(ServicePoint, id=service_point_id)

    if request.method == 'POST':
        patient_info = {
            'patient_type': request.POST.get('patient_type'),
            'user_role': request.POST.get('user_role'),
            'benefit_plan': request.POST.get('benefit_plan'),
            'benefit_plan_other': request.POST.get('benefit_plan_other', ''), 
            'age_range': request.POST.get('age_range'),
            'gender': request.POST.get('patient_gender', 'NOT_SPECIFIED'),
        }
        request.session['patient_info'] = patient_info
        return redirect('survey:survey_display', service_point_id=service_point.id)
    
    # [แก้ไข] สร้างรายการตัวเลือกที่นี่ แล้วส่งไปที่ Template
    age_ranges = ["ต่ำกว่า 20 ปี", "20-39 ปี", "40-59 ปี", "60 ปีขึ้นไป"]

    context = { 
        'service_point': service_point,
        'age_ranges': Response.AgeRange.choices,# ส่งค่าไป
    }
    return render(request, 'kiosk/kiosk_user_info.html', context)

def kiosk_thank_you_view(request, service_point_id):
    context = {'service_point_id': service_point_id}
    return render(request, 'kiosk/kiosk_thank_you.html', context)

def survey_display_view(request, service_point_id):
    service_point = get_object_or_404(ServicePoint, id=service_point_id)
    
    print(f"--- DEBUG DISPLAY VIEW (SP: {service_point.id}) ---")
    
    # 1. ลองหาตัวที่ Active ก่อน (ตามที่คุณต้องการ)
    active_survey = Survey.objects.filter(
        service_point=service_point,
        status=Survey.Status.ACTIVE
    ).order_by('-id').first()

    if active_survey:
        print(f"✅ เจอแบบสอบถาม ACTIVE: ID {active_survey.id} ({active_survey.title_th})")
    else:
        print("❌ ไม่เจอแบบสอบถาม ACTIVE... กำลังลองหาตัวล่าสุดที่เป็น DRAFT แทน...")
        
        # 2. ถ้าไม่เจอ (Fallback) -> เอาตัวล่าสุดมาเลย (กันหน้าจอขาว)
        active_survey = Survey.objects.filter(
            service_point=service_point
        ).order_by('-id').first()
        
        if active_survey:
            print(f"⚠️ เจอตัวสำรอง (Status: {active_survey.status}): ID {active_survey.id}")
        else:
            print("💀 ไม่เจออะไรเลยในจุดบริการนี้")

    if not active_survey:
        # ถ้าหาไม่เจอจริงๆ ค่อยยอมแพ้
        return render(request, 'kiosk/survey_display.html', {
            'service_point': service_point,
            'survey': None
        })
    
    context = {
        'service_point': service_point,
        'survey': active_survey,
    }
    return render(request, 'kiosk/survey_display.html', context)

def survey_submit_view(request, survey_id):
    if request.method != 'POST':
        return redirect('survey:kiosk_welcome', service_point_id=1) 
    
    survey = get_object_or_404(Survey, id=survey_id) 
    service_point_id = request.POST.get('service_point_id')
    service_point = get_object_or_404(ServicePoint, id=service_point_id)
    patient_info = request.session.get('patient_info', {})

    # 1. บันทึก Response
    response = Response.objects.create(
        survey=survey, 
        service_point=service_point,
        patient_type=patient_info.get('patient_type'),
        user_role=patient_info.get('user_role'),
        benefit_plan=patient_info.get('benefit_plan'),
        benefit_plan_other=patient_info.get('benefit_plan_other'),
        age_range=patient_info.get('age_range'),
        gender=patient_info.get('gender'),
        submitted_at=timezone.localtime(timezone.now())
    )

    # 2. บันทึก Answers
    for key, value in request.POST.items():
        if key.startswith('q-'):
            if not value: continue
            try:
                question_id = key.split('-')[1]
                question = Question.objects.get(id=question_id)
                
                if question.question_type == 'RATING_5':
                    ResponseAnswer.objects.create(
                        response=response, 
                        question=question, 
                        answer_rating=int(value)
                    )
                elif question.question_type == 'TEXTAREA':
                    ResponseAnswer.objects.create(
                        response=response, 
                        question=question, 
                        answer_text=str(value)
                    )
            except (Question.DoesNotExist, ValueError): 
                continue 
    
    # ====================================================
    # 🔔 3. LOGIC แจ้งเตือนเมื่อคะแนนต่ำ (< 2.5)
    # ====================================================
    
    avg_score = response.answers.aggregate(avg=Avg('answer_rating'))['avg'] or 0
    FULL_DOMAIN = settings.FULL_DOMAIN # ดึงค่า Domain จาก settings.py

    # ตั้งเกณฑ์ที่ต้องการแจ้งเตือน (น้อยกว่า 2.5)
    if avg_score > 0 and avg_score < 2.5:
        
        # A. หา Manager และ Admin ทั้งหมด
        managers = list(User.objects.filter(managed_points=service_point))
        admins = list(User.objects.filter(is_superuser=True))
        recipients = set(managers + admins)
        
        # 📌 NEW: ดึงรายชื่ออีเมล Manager ทั้งหมด (สำหรับส่งอีเมลรวม)
        manager_emails = [m.email for m in managers if m.email] 
        
        # ข้อความและลิงก์พื้นฐาน
        line_title = f"⚠️ คะแนนต่ำผิดปกติ ({avg_score:.1f})"
        line_message_base = f"จุดบริการ: {service_point.name}\nผู้ประเมิน: {response.user_role or 'ไม่ระบุ'}"
        
        
        for user in recipients:
            
            # กำหนดลิงก์ปลายทางที่แตกต่างกัน
            if user.is_superuser:
                # ลิงก์สำหรับ Admin Portal
                link = f"/survey/assessments/?survey_id={survey.id}&point_id={service_point.id}&popup=true"
            else:
                # ลิงก์สำหรับ Manager Dashboard/Response
                link = f"/manager/response/?point_id={service_point.id}&score=1-2&popup=true"

            # 1. สร้าง Notification ใน Database
            Notification.objects.create(
                recipient=user,
                title=line_title,
                message=line_message_base,
                link=link
            )

            # 2. ส่ง LINE แจ้งเตือน Admin/Manager รายบุคคล (ถ้ามี Line ID)
            try:
                line_id = user.profile.line_user_id
                if line_id:
                    full_link = f"{FULL_DOMAIN}{link}"
                    line_message = f"🚨 [แจ้งเตือนบุคคล]\n{line_title}\n{line_message_base}\n\nตรวจสอบ: {full_link}"
                    send_line_push(line_message, line_id)
            except UserProfile.DoesNotExist:
                # ถ้าผู้ใช้ไม่มี Profile หรือ Line ID ให้ข้ามการส่ง LINE
                pass
                
        # 3. ส่งแจ้งเตือนไปหา Admin กลาง (LINE)
        admin_link = f"{FULL_DOMAIN}/survey/assessments/?survey_id={survey.id}&point_id={service_point.id}"
        admin_line_alert = f"📢 [แจ้งเตือน Admin Portal]\n{line_title}\n{line_message_base}\n\nตรวจสอบ: {admin_link}"
        
        # สมมติว่า LINE_ADMIN_RECIPIENT_ID เป็น UID ของผู้รับคนกลางใน settings.py
        send_line_push(admin_line_alert, settings.LINE_ADMIN_RECIPIENT_ID)


        # ====================================================
        # 📌 4. NEW: ส่ง EMAIL แจ้งเตือน Manager (ใช้รายชื่ออีเมลที่ดึงมาก่อนหน้านี้)
        # ====================================================
        if manager_emails:
            email_subject = f"[ALERT] Survey Hospital: คะแนนต่ำ ({avg_score:.1f}) ที่ {service_point.name}"
            email_body = (
                f"เรียน ผู้จัดการทุกท่าน,\n\n"
                f"ตรวจพบการประเมินคะแนนความพึงพอใจต่ำกว่า 2.5 คะแนน\n\n"
                f"หัวข้อ: {line_title}\n"
                f"คะแนนเฉลี่ยที่ได้: {avg_score:.1f} / 5.0\n"
                f"จุดบริการ: {service_point.name}\n"
                f"ผู้ประเมิน: {response.user_role or 'ไม่ระบุ'}\n"
                f"เวลาที่ส่ง: {response.submitted_at.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                f"โปรดเข้าสู่ระบบเพื่อตรวจสอบรายละเอียด: {FULL_DOMAIN}/manager/response/?point_id={service_point.id}&score=1-2"
            )
            
            send_email_alert(email_subject, email_body, manager_emails)
        else:
            print(f"🚨 ไม่พบ Manager ที่รับผิดชอบจุดบริการ {service_point.name} หรือไม่มีอีเมลสำหรับส่ง")
        
        # ====================================================

    # ====================================================

    if 'patient_info' in request.session:
        del request.session['patient_info']

    return redirect('survey:kiosk_thank_you', service_point_id=service_point_id)

# --- Export Views ---

def _get_base_filtered_responses(request):
    """ฟังก์ชันกลางสำหรับกรอง Response ตาม User, Date, Group, Point"""
    user = request.user
    
    # 1.1 Permission Filter
    base_service_points = ServicePoint.objects.all()
    if user.is_authenticated and not user.is_superuser:
        managed_points = user.managed_points.all()
        base_service_points = base_service_points.filter(id__in=managed_points.values('id'))

    # 1.2 Date Filter (Logic เดียวกับหน้าเว็บ)
    default_end = timezone.now().date()
    default_start = (timezone.now() - timedelta(days=30)).date()
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else default_end
    except (ValueError, TypeError):
        start_date = default_start
        end_date = default_end

    end_date_query = end_date + timedelta(days=1)

    # 1.3 Query Responses
    responses = Response.objects.filter(
        service_point__in=base_service_points,
        submitted_at__gte=start_date,
        submitted_at__lt=end_date_query
    )

    # 1.4 Group & Point Filter
    group_id = request.GET.get('group_id')
    point_id = request.GET.get('point_id')

    if group_id and group_id.isdigit():
        responses = responses.filter(service_point__group_id=int(group_id))
    if point_id and point_id.isdigit():
        responses = responses.filter(service_point_id=int(point_id))
        
    return responses

def export_assessment_excel(request):
    # 1. ใช้ Helper ดึงข้อมูลที่กรองแล้ว
    responses = _get_base_filtered_responses(request)
    
    # 2. ดึงคำตอบทั้งหมด (ทั้งคะแนนและข้อความ)
    queryset = ResponseAnswer.objects.filter(response__in=responses)\
        .select_related('response', 'response__service_point', 'question')\
        .order_by('response__submitted_at')

    # 3. สร้าง Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Data"
    
    headers = ['Response ID', 'Service Point', 'Submitted At', 'Role', 'Question', 'Answer (Rating/Text)']
    ws.append(headers)
    
    # จัดความกว้าง
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20

    for ans in queryset:
        local_time = timezone.localtime(ans.response.submitted_at).replace(tzinfo=None)
        
        # เลือกค่าที่จะแสดง (คะแนน หรือ ข้อความ)
        val = ans.answer_rating if ans.answer_rating is not None else ans.answer_text
        
        ws.append([
            ans.response.id,
            ans.response.service_point.name,
            local_time,
            ans.response.user_role,
            getattr(ans.question, 'text_th', ''),
            val
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="assessment_data.xlsx"'
    wb.save(response)
    return response

# ==========================================
# PART B: Export สำหรับหน้า "ข้อเสนอแนะ" (เฉพาะ Comment)
# ==========================================

def export_suggestion_excel(request):
    # 1. ใช้ Helper ดึงข้อมูลชุดเดียวกัน แต่กรองเฉพาะ Text
    responses = _get_base_filtered_responses(request)
    
    # 2. กรองเอาเฉพาะข้อเสนอแนะ (TEXTAREA) และไม่เอาค่าว่าง
    queryset = ResponseAnswer.objects.filter(
        response__in=responses,
        question__question_type='TEXTAREA'
    ).exclude(answer_text='')\
    .select_related('response', 'response__service_point')\
    .order_by('response__submitted_at')

    # 3. สร้าง Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suggestions"
    
    headers = ['Date/Time', 'Service Point', 'Group', 'User Role', 'Suggestion']
    ws.append(headers)
    
    # จัดความกว้าง
    ws.column_dimensions['A'].width = 22 # Date
    ws.column_dimensions['B'].width = 30 # Point
    ws.column_dimensions['C'].width = 25 # Group
    ws.column_dimensions['E'].width = 60 # Suggestion (กว้างหน่อย)

    for ans in queryset:
        local_time = timezone.localtime(ans.response.submitted_at).replace(tzinfo=None)
        
        ws.append([
            local_time,
            ans.response.service_point.name,
            ans.response.service_point.group.name,
            ans.response.user_role,
            ans.answer_text
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="suggestion_list.xlsx"'
    wb.save(response)
    return response

def export_assessment_csv(request):
    # 1. ใช้ Helper ตัวเดิม (เพื่อให้ Filter ตรงกับหน้าเว็บ)
    responses = _get_base_filtered_responses(request)
    
    # 2. Query Data
    queryset = ResponseAnswer.objects.filter(response__in=responses)\
        .select_related('response', 'response__service_point', 'question')\
        .order_by('response__submitted_at')

    # 3. เตรียมไฟล์ CSV
    response = HttpResponse(content_type='text/csv', headers={'Content-Disposition': 'attachment; filename="assessment_data.csv"'})
    response.write('\ufeff') # สำคัญมาก! เพื่อให้ Excel อ่านภาษาไทยออก
    
    writer = csv.writer(response)
    # Header
    writer.writerow(['Response ID', 'Service Point', 'Submitted At', 'Role', 'Question', 'Answer (Rating/Text)'])
    
    # Loop Data
    for ans in queryset:
        # เลือกค่าที่จะแสดง
        val = ans.answer_rating if ans.answer_rating is not None else ans.answer_text
        
        writer.writerow([
            ans.response.id,
            ans.response.service_point.name,
            ans.response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            ans.response.user_role,
            getattr(ans.question, 'text_th', ''),
            val
        ])
    return response

def export_suggestion_csv(request):
    # 1. ใช้ Helper ตัวเดิม
    responses = _get_base_filtered_responses(request)
    
    # 2. Query เฉพาะข้อเสนอแนะ
    queryset = ResponseAnswer.objects.filter(
        response__in=responses,
        question__question_type='TEXTAREA'
    ).exclude(answer_text='')\
    .select_related('response', 'response__service_point', 'response__service_point__group')\
    .order_by('response__submitted_at')

    # 3. เตรียมไฟล์ CSV
    response = HttpResponse(content_type='text/csv', headers={'Content-Disposition': 'attachment; filename="suggestion_list.csv"'})
    response.write('\ufeff') # BOM สำหรับภาษาไทย
    
    writer = csv.writer(response)
    # Header
    writer.writerow(['Date/Time', 'Service Point', 'Group', 'User Role', 'Suggestion'])
    
    # Loop Data
    for ans in queryset:
        writer.writerow([
            ans.response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            ans.response.service_point.name,
            ans.response.service_point.group.name,
            ans.response.user_role,
            ans.answer_text
        ])
    return response


def export_dashboard_summary(request):
    """
    ฟังก์ชัน Export Dashboard Summary (ฉบับแก้ไข: ดึง User ผู้ดูแลทุกคน ไม่จำกัดแค่ Staff)
    """
    
    # 1. จัดการวันที่ (Header)
    default_end = timezone.now().date()
    default_start = (timezone.now() - timedelta(days=30)).date()
    
    req_start = request.GET.get('start_date')
    req_end = request.GET.get('end_date')

    show_start_date = req_start if req_start else default_start.strftime('%Y-%m-%d')
    show_end_date = req_end if req_end else default_end.strftime('%Y-%m-%d')

    # 2. ดึงข้อมูล Response (Filtered)
    responses = _get_base_filtered_responses(request)

    # --- KPI ---
    total_responses = responses.count()
    total_service_points = ServicePoint.objects.count()
    active_questions = Question.objects.count()

    # --- Top Service Points ---
    sp_stats = responses.values('service_point__name')\
        .annotate(total=Count('id'))\
        .order_by('-total')

    # --- Weekly Stats (Pure Python) ---
    raw_datetimes = responses.values_list('submitted_at', flat=True)
    weekly_stats = {}
    
    for dt in raw_datetimes:
        if dt is None: continue
        local_date = timezone.localtime(dt).date()
        monday = local_date - timedelta(days=local_date.weekday())
        if monday not in weekly_stats:
            weekly_stats[monday] = 0
        weekly_stats[monday] += 1

    weekly_stats_list = [{'week': k, 'total': v} for k, v in weekly_stats.items()]
    weekly_stats_list.sort(key=lambda x: x['week'])

    # --- [จุดที่แก้ไข] รายชื่อผู้ดูแล ---
    # เปลี่ยนเงื่อนไข: เอาเฉพาะ User ที่มี managed_points (อยู่ในตาราง servicepoint_manager)
    admins = User.objects.filter(is_active=True, managed_points__isnull=False)\
        .distinct()\
        .prefetch_related('managed_points')

    # 3. สร้าง Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Summary"

    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)

    # Header
    ws.append(["รายงานสรุปผลการประเมิน (Dashboard Summary)"])
    ws['A1'].font = header_font
    ws.append([f"ช่วงเวลาข้อมูล: {show_start_date} ถึง {show_end_date}"])
    ws.append([]) 

    # Section 1
    ws.append(["1. ภาพรวม (KPIs)"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.append(["หัวข้อ", "จำนวน"])
    ws.append(["จำนวนการตอบทั้งหมด", total_responses])
    ws.append(["จำนวนจุดบริการทั้งหมด", total_service_points])
    ws.append(["จำนวนคำถามทั้งหมด", active_questions])
    ws.append([])

    # Section 2
    ws.append(["2. จำนวนการประเมินแยกตามจุดบริการ"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.append(["ชื่อจุดบริการ", "จำนวนครั้ง"])
    for item in sp_stats:
        ws.append([item['service_point__name'], item['total']])
    ws.append([])

    # Section 3
    ws.append(["3. สถิติรายสัปดาห์ (Weekly Trend)"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.append(["สัปดาห์ (เริ่มวันจันทร์)", "จำนวนการประเมิน"])
    if not weekly_stats_list:
        ws.append(["ไม่พบข้อมูลในช่วงเวลานี้", "-"])
    else:
        for item in weekly_stats_list:
            week_str = item['week'].strftime('%Y-%m-%d')
            ws.append([week_str, item['total']])
    ws.append([])

    # Section 4
    ws.append(["4. รายชื่อผู้ดูแลและจุดบริการที่รับผิดชอบ"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.append(["ชื่อผู้ดูแล", "จุดบริการที่ดูแล"])
    
    for admin in admins:
        # ดึงชื่อจุดบริการทั้งหมดที่ User คนนี้ดูแล
        points_list = [p.name for p in admin.managed_points.all()]
        points_str = ", ".join(points_list)
        
        admin_name = admin.get_full_name()
        if not admin_name:
            admin_name = admin.username
            
        ws.append([admin_name, points_str])

    # Styling
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 60

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dashboard_summary.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_superuser) # หรือเช็คตามสิทธิ์ของคุณ
def assessment_results_view(request):
    user = request.user

    # 1. รับค่า Filter จาก URL
    group_id = request.GET.get('group_id')
    point_id = request.GET.get('point_id')
    score_filter = request.GET.get('score')
    
    # --- [ส่วนสำคัญ 1] จัดการวันที่ให้เหมือน Export เป๊ะๆ ---
    end_date_str = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    start_date_str = request.GET.get('start_date', (timezone.now() - timedelta(days=6)).strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        # ถ้าวันที่เพี้ยน ให้กลับไปใช้ค่า Default (7 วันล่าสุด)
        start_date = (timezone.now() - timedelta(days=6)).date()
        end_date = timezone.now().date()
        # อัปเดต string ให้ตรงกับ date ที่ใช้จริง (เพื่อส่งกลับไปหน้าเว็บ)
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

    end_date_for_query = end_date + timedelta(days=1)
    # ----------------------------------------------------

    # 2. Base Query (กรองตามสิทธิ์ก่อน)
    base_service_points = ServicePoint.objects.all()
    if user.is_authenticated and not user.is_superuser:
        managed_points = user.managed_points.all()
        base_service_points = base_service_points.filter(id__in=managed_points.values('id'))

    # 3. สร้าง Queryset หลัก
    responses = Response.objects.filter(
        service_point__in=base_service_points,
        submitted_at__gte=start_date,
        submitted_at__lt=end_date_for_query
    ).annotate(
        avg_score=Avg('answers__answer_rating')
    ).select_related('service_point', 'service_point__group').order_by('-submitted_at')

    # 4. Apply Filters (กรองเพิ่มเติมตามที่เลือก)
    if group_id:
        responses = responses.filter(service_point__group_id=group_id)
    
    if point_id:
        responses = responses.filter(service_point_id=point_id)

    if score_filter:
        try:
            min_score, max_score = map(int, score_filter.split('-'))
            if max_score == 5:
                responses = responses.filter(avg_score__gte=min_score, avg_score__lte=max_score)
            else:
                responses = responses.filter(avg_score__gte=min_score, avg_score__lt=max_score)
        except ValueError:
            pass

    # 5. Stats & Pagination (ส่วนแสดงผล)
    total_assessments = responses.count()
    
    # หา Suggestion (ต้อง filter ตาม responses ที่กรองมาแล้ว)
    suggestion_queryset = ResponseAnswer.objects.filter(
        response__in=responses,
        question__question_type='TEXTAREA'
    ).exclude(answer_text='')
    
    total_suggestions = suggestion_queryset.count()
    recent_suggestions = suggestion_queryset.select_related('response', 'response__service_point').order_by('-id')[:10]

    paginator = Paginator(responses, 10) 
    page_obj = paginator.get_page(request.GET.get('page'))

    # 6. Prepare Context
    groups = ServiceGroup.objects.all()
    points = ServicePoint.objects.all() # ส่งไปทั้งหมดเพื่อให้ JS จัดการ หรือจะกรองก็ได้
    
    context = {
        'page_title': 'ผลการประเมิน',
        'total_assessments': total_assessments,
        'total_suggestions': total_suggestions,
        'page_obj': page_obj,
        'recent_suggestions': recent_suggestions,
        'groups': groups,
        'points': points,
        'point_map_json': json.dumps(_get_point_map()), 
        # ส่งค่ากลับไปที่ Form (สำคัญมาก: ต้องส่งค่าที่ใช้จริงกลับไป)
        'selected_group': int(group_id) if group_id and group_id.isdigit() else '',
        'selected_point': int(point_id) if point_id and point_id.isdigit() else '',
        'selected_score': score_filter,
        'start_date': start_date_str, # ส่ง string กลับไปให้ input date
        'end_date': end_date_str,     # ส่ง string กลับไปให้ input date
    }
    return render(request, 'survey/assessment_results.html', context)

def _get_base_response_queryset(user, start_date, end_date):
    """
    ฟังก์ชันภายใน: ดึงข้อมูล Response ตามสิทธิ์ User และช่วงเวลาเท่านั้น
    """
    # 1. กรองตามสิทธิ์ (User Permissions)
    base_service_points = ServicePoint.objects.all()
    if user.is_authenticated and not user.is_superuser:
        managed_points = user.managed_points.all()
        base_service_points = base_service_points.filter(id__in=managed_points.values('id'))
    
    # 2. Query พื้นฐาน
    end_date_for_query = end_date + timedelta(days=1)
    return Response.objects.filter(
        service_point__in=base_service_points,
        submitted_at__gte=start_date,
        submitted_at__lt=end_date_for_query
    )


@login_required
@user_passes_test(is_superuser)
def suggestion_list_view(request):
    # 1. รับค่า Filter
    group_id = request.GET.get('group_id')
    point_id = request.GET.get('point_id')
    search_query = request.GET.get('q', '')
    
    # --- ส่วนวันที่ (เหมือนเดิม) ---
    default_end = timezone.now().date()
    default_start = (timezone.now() - timedelta(days=30)).date()
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else default_end
    except (ValueError, TypeError):
        start_date = default_start
        end_date = default_end

    # 2. Base Query
    suggestions = ResponseAnswer.objects.filter(
        question__question_type='TEXTAREA'
    ).exclude(answer_text='').select_related(
        'response', 'response__service_point', 'response__service_point__group'
    ).order_by('-response__submitted_at')
    
    # 3. Apply Filters
    end_date_query = end_date + timedelta(days=1)
    suggestions = suggestions.filter(response__submitted_at__gte=start_date, response__submitted_at__lt=end_date_query)

    if group_id:
        suggestions = suggestions.filter(response__service_point__group_id=group_id)
    if point_id:
        suggestions = suggestions.filter(response__service_point_id=point_id)
    if search_query:
        suggestions = suggestions.filter(answer_text__icontains=search_query)

    # 4. Pagination (หน้าละ 10 รายการ)
    paginator = Paginator(suggestions, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    # 5. Context
    groups = ServiceGroup.objects.all()
    # ไม่ต้อง query points ทั้งหมดแล้ว เพราะจะใช้ JSON map แทน
    
    context = {
        'page_title': 'รายการข้อเสนอแนะทั้งหมด',
        'page_obj': page_obj,
        'groups': groups,
        'point_map_json': json.dumps(_get_point_map()), # [สำคัญ] ส่ง Map ไปให้ JS
        'selected_group': int(group_id) if group_id and group_id.isdigit() else '',
        'selected_point': int(point_id) if point_id and point_id.isdigit() else '',
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'search_query': search_query,
    }
    return render(request, 'survey/suggestion_list.html', context)

@login_required
def check_notifications(request):
    unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()
    notifs_qs = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    show_all = request.GET.get('all') == 'true'
    
    if show_all:
        latest_notifs = notifs_qs[:50]
    else:
        latest_notifs = notifs_qs[:5]

    has_more = notifs_qs.count() > 5

    notif_list = []
    for n in latest_notifs:
        notif_list.append({
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'read_url': reverse('survey:read_notification', args=[n.id]), 
            'time_ago': timesince(n.created_at) + " ที่แล้ว",
            'is_read': n.is_read
        })
    
    return JsonResponse({
        'unread_count': unread_count,
        'notifications': notif_list,
        'has_more': has_more 
    })

@login_required
def mark_notification_read(request, notif_id):

    notification = get_object_or_404(Notification, pk=notif_id, recipient=request.user)

    if not notification.is_read:
        notification.is_read = True
        notification.save()

    return redirect(notification.link if notification.link else 'manager:dashboard')


from django.views.decorators.http import require_POST
@login_required
@require_POST
def clear_all_notifications(request):
    """ลบการแจ้งเตือนทั้งหมด (สำหรับ Admin)"""
    Notification.objects.filter(recipient=request.user).delete()
    return JsonResponse({'status': 'success'})


