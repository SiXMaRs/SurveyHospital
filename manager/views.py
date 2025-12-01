from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.sessions.models import Session
from django.views.generic import UpdateView, DeleteView
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator
from django.contrib import messages
from datetime import datetime, timedelta
from django.urls import reverse
from django.db.models import Count, Q, Avg, F
from survey.models import *
from .forms import * 
import json
import csv
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

@login_required
def dashboard_view(request):
    user = request.user
    base_service_points = user.managed_points.all()
    # Manager List: แสดงแค่ตัว Manager เองคนเดียว (เพื่อให้ Template ไม่ต้องแก้เยอะ)
    managers_list = User.objects.filter(
        managed_points__in=base_service_points
    ).distinct().prefetch_related('managed_points') 
    
    today = timezone.now().date()
    # Default: สัปดาห์ปัจจุบัน (จันทร์ - อาทิตย์)
    start_date_default = today - timedelta(days=today.weekday()) 
    end_date_default = start_date_default + timedelta(days=6) 
    end_date_str = request.GET.get('end_date', end_date_default.strftime('%Y-%m-%d'))
    start_date_str = request.GET.get('start_date', start_date_default.strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = start_date_default
        end_date = end_date_default
        
    # บวก 1 วันเพื่อให้ครอบคลุมเวลา 23:59:59 ของวันสิ้นสุด
    end_date_for_query = end_date + timedelta(days=1)
    
    # --- 3. สร้าง Queryset หลัก (Main Filter) ---
    # กรอง Response ทั้งหมด โดยต้องมาจากจุดบริการของเรา (base_service_points) และอยู่ในช่วงเวลา
    filtered_responses = Response.objects.filter(
        service_point__in=base_service_points, 
        submitted_at__gte=start_date, 
        submitted_at__lt=end_date_for_query 
    )

    # --- 4. คำนวณข้อมูล (KPIs) ---
    # จำนวนคำถามที่ Active (เท่ากันทุกคน)
    total_active_questions = Question.objects.filter(
        survey__status=Survey.Status.ACTIVE
    ).count()
    
    # ยอดรวมการตอบ (เฉพาะของ Manager คนนี้)
    total_responses = filtered_responses.count()
    total_service_points_in_view = base_service_points.count()
    total_managers = managers_list.count()

    # --- 5. List: รายการจุดบริการพร้อมยอด (Service Point List) ---
    all_service_points_with_counts = base_service_points.annotate(
        response_count=Count('response', filter=Q(response__in=filtered_responses))
    ).order_by('-response_count')

    # --- 6. Chart: กราฟแท่ง (Bar Chart) ---
    date_labels = []
    day_counts_dict = {}
    current_date = start_date
    
    # สร้างแกน X ให้ครบทุกวันในช่วงที่เลือก
    while current_date <= end_date:
        date_labels.append(current_date.strftime('%a %d/%m')) 
        day_counts_dict[current_date] = 0
        current_date += timedelta(days=1)

    # เติมข้อมูลลงในวันที่มีการตอบ
    response_times = filtered_responses.values_list('submitted_at', flat=True)
    for submitted_at_utc in response_times:
        local_time = timezone.localtime(submitted_at_utc)
        date_only = local_time.date() 
        if date_only in day_counts_dict:
            day_counts_dict[date_only] += 1
            
    bar_data_weekly = [day_counts_dict[day] for day in sorted(day_counts_dict.keys())]

    # --- 7. Chart: กราฟวงกลม (Pie Chart) ---
    pie_data_query = all_service_points_with_counts.filter(response_count__gt=0)
    pie_labels = [sp.name for sp in pie_data_query]
    pie_data = [sp.response_count for sp in pie_data_query]

    # --- 8. ข้อเสนอแนะ (Feedback Ticker) ---
    recent_feedback = ResponseAnswer.objects.filter(
        response__in=filtered_responses, # กรองแล้ว
        question__question_type='TEXTAREA'
    ).exclude(
        Q(answer_text__isnull=True) | Q(answer_text__exact='')
    ).select_related(
        'response__service_point'
    ).order_by('-response__submitted_at')[:5]
    
    # --- 9. ส่ง Context ไป Template ---
    context = {
        'total_responses': total_responses,
        'total_managers': total_managers,
        'total_active_questions': total_active_questions,
        
        'all_service_points_with_counts': all_service_points_with_counts,
        
        'bar_labels_weekly': json.dumps(date_labels),
        'bar_data_weekly': json.dumps(bar_data_weekly),
        
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
        
        'managers_list': managers_list,
        'recent_feedback': recent_feedback,
        
        'start_date': start_date_str,
        'end_date': end_date_str,
    }

    return render(request, 'manager/dashboard.html', context)

@login_required
def manager_list_view(request):
    user = request.user
    
    # 1. หาจุดบริการและเพื่อนร่วมงาน (co_managers)
    my_points = user.managed_points.all()
    co_managers = User.objects.filter(
        managed_points__in=my_points,
        is_superuser=False
    ).distinct().prefetch_related('managed_points').order_by('username')

    # 2. Search Logic (ถ้ามี)
    query = request.GET.get('q', '')
    if query:
        co_managers = co_managers.filter(
            Q(username__icontains=query) | Q(email__icontains=query)
        )

    # 3. Online Status Logic (หา Session ทั้งหมดที่ Active)
    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    all_online_user_ids = set()
    for session in sessions:
        data = session.get_decoded()
        uid = data.get('_auth_user_id')
        if uid:
            all_online_user_ids.add(int(uid))

    co_manager_ids = set(co_managers.values_list('id', flat=True))
    
    # หาจุดตัด (คนที่ออนไลน์ AND เป็นเพื่อนร่วมงานเรา)
    online_co_managers_count = len(co_manager_ids.intersection(all_online_user_ids))

    context = {
        'page_title': 'ผู้ดูแลจุดบริการ (ทีมงาน)',
        'managers': co_managers,
        'search_query': query,
        'online_user_ids': all_online_user_ids, # ส่งไปใช้ใน loop ตารางเหมือนเดิม
        'online_count': online_co_managers_count, # 🔴 ส่งตัวเลขที่ถูกต้องไปแสดงในการ์ด
        'my_points': my_points,
    }
    
    return render(request, 'manager/manager_list.html', context)

@login_required
def survey_list_view(request):
    user = request.user
    my_points = user.managed_points.all()

    surveys = Survey.objects.filter(
        service_point__in=my_points
    ).select_related('service_point', 'service_point__group').annotate(
        question_count=Count('questions') 
    ).order_by('-created_at')

    # --- 📊 ส่วนที่ 1: คำนวณ Stats สำหรับการ์ด (คำนวณจากทั้งหมดก่อนกรอง) ---
    total_surveys = surveys.count()
    total_questions = Question.objects.filter(survey__in=surveys).count()
    total_service_points = my_points.count()

    # --- 🔍 ส่วนที่ 2: Logic การกรอง (Filter Bar) ---
    search_query = request.GET.get('q')
    group_filter = request.GET.get('group')
    point_filter = request.GET.get('point')

    if search_query:
        surveys = surveys.filter(
            Q(title_th__icontains=search_query) | 
            Q(title_en__icontains=search_query)
        )
    
    if group_filter:
        surveys = surveys.filter(service_point__group_id=group_filter)
    
    if point_filter:
        surveys = surveys.filter(service_point_id=point_filter)

    # เตรียมข้อมูลใส่ Dropdown ใน Filter Bar (เฉพาะที่เกี่ยวข้องกับ Manager)
    filter_groups = ServiceGroup.objects.filter(service_points__in=my_points).distinct()
    filter_points = my_points

    # --- 📝 ส่วนที่ 3: Handle Form (Create New) ---
    show_modal = False 

    if request.method == 'POST':
        form = ManagerSurveyForm(user, request.POST)
        if form.is_valid():
            new_status = form.cleaned_data.get('status')
            new_service_point = form.cleaned_data.get('service_point')

            # === CONSTRAINT CHECK (1 Active per Point) ===
            if new_status == 'ACTIVE':
                if Survey.objects.filter(
                    service_point=new_service_point,
                    status='ACTIVE'
                ).exists():
                    messages.error(request, f"ไม่สามารถเปิดใช้งานได้: จุดบริการ **'{new_service_point.name}'** มีแบบสอบถามที่เปิดใช้งานอยู่แล้ว")
                    show_modal = True
                    
                    # ส่ง Context กลับไปเพื่อให้หน้าเว็บไม่พังและ Modal เปิดค้างไว้
                    context = {
                        'page_title': 'จัดการแบบสอบถาม',
                        'surveys': surveys,
                        'form': form,
                        'show_modal': show_modal,
                        # ต้องส่งค่าพวกนี้กลับไปด้วย ไม่งั้นหน้าจอจะ error หรือการ์ดหาย
                        'total_surveys': total_surveys,
                        'total_questions': total_questions,
                        'total_service_points': total_service_points,
                        'groups': filter_groups,
                        'service_points': filter_points,
                    }
                    return render(request, 'manager/survey_list.html', context)

            # ถ้าผ่านเงื่อนไข -> บันทึก
            survey = form.save(commit=False)
            survey.version_number = "1.0"
            survey.save()
            messages.success(request, "สร้างแบบสอบถามเรียบร้อยแล้ว")
            return redirect('manager:survey_list')
        else:
            show_modal = True
            messages.error(request, "เกิดข้อผิดพลาด กรุณาตรวจสอบข้อมูล")
    else:
        form = ManagerSurveyForm(user)

    # --- 📦 ส่วนที่ 4: Context Final ---
    context = {
        'page_title': 'จัดการแบบสอบถาม',
        'surveys': surveys,
        'form': form,
        'show_modal': show_modal,
        
        # Stats Cards Variables
        'total_surveys': total_surveys,
        'total_questions': total_questions,
        'total_service_points': total_service_points,
        
        # Dropdown Choices
        'groups': filter_groups,
        'service_points': filter_points,
    }
    return render(request, 'manager/survey_list.html', context)
    
@login_required
def survey_edit_view(request, pk):
    # ต้องดึง survey เดิมมาก่อน
    original_survey = get_object_or_404(Survey, pk=pk, service_point__in=request.user.managed_points.all())

    if request.method == 'POST':
        # 🔴 ใช้ original_survey เป็น instance ใน form
        form = ManagerSurveyForm(request.user, request.POST, instance=original_survey)
        
        if form.is_valid():
            changed_data = form.changed_data
            new_status = form.cleaned_data.get('status')
            new_service_point = form.cleaned_data.get('service_point') # ดึง Service Point ใหม่

            # --- 🔴 STEP 1: CONSTRAINT CHECK (1 ACTIVE SURVEY PER SERVICE POINT) ---
            if new_status == 'ACTIVE':
                # ตรวจสอบว่ามีแบบสอบถามอื่น (ที่ไม่ใช่ตัวปัจจุบัน) ที่ Active อยู่แล้วหรือไม่
                if Survey.objects.filter(
                    service_point=new_service_point,
                    status='ACTIVE'
                ).exclude(pk=original_survey.pk).exists():
                    
                    messages.error(request, f"ไม่สามารถเปิดใช้งานได้: จุดบริการ **'{new_service_point.name}'** มีแบบสอบถามที่เปิดใช้งานอยู่แล้ว")
                    return redirect('manager:survey_list')

            # --- STEP 2: APPLY SAVE LOGIC (Versioning) ---
            
            # CASE A: Change only Status -> Update In-place (ไม่ต้องสร้างเวอร์ชันใหม่)
            if len(changed_data) == 1 and 'status' in changed_data:
                # ถ้าเปลี่ยนสถานะอย่างเดียว (และผ่าน Constraint Check แล้ว)
                form.save()
            else:
                try:
                    with transaction.atomic():
                        
                        # 2.1 [Cleanup] ถ้าเวอร์ชันใหม่เป็น ACTIVE, ควรกำหนดให้เวอร์ชันเดิมเป็น DRAFT
                        if new_status == 'ACTIVE':
                             Survey.objects.filter(pk=original_survey.pk).update(status='DRAFT')

                        # 2.2 สร้าง Survey Object ใหม่
                        new_survey = form.save(commit=False)
                        new_survey.pk = None # สำคัญ: ลบ PK เพื่อให้ Django สร้าง record ใหม่
                        
                        # Version logic: เพิ่มเลขเวอร์ชัน
                        try:
                            current_ver = float(original_survey.version_number or 0)
                        except ValueError:
                            current_ver = 0 # กรณีที่ format ผิด
                            
                        new_survey.version_number = f"{int(current_ver) + 1}.0"
                        
                        new_survey.save()

                        # 2.3 Clone Questions (คัดลอกคำถามจากเวอร์ชันเดิม)
                        old_questions = original_survey.questions.all().order_by('order')
                        new_questions = [
                            Question(
                                survey=new_survey,
                                # **ใส่ฟิลด์คำถามทั้งหมดที่ต้องการโคลนมาที่นี่**
                                text_th=q.text_th,
                                text_en=q.text_en,
                                question_type=q.question_type,
                                order=q.order,
                                is_required=q.is_required
                                # ... เพิ่มฟิลด์อื่น ๆ เช่น choices, min_value, max_value ตาม Model ...
                            ) for q in old_questions
                        ]
                        Question.objects.bulk_create(new_questions)

                except Exception as e:
                    messages.error(request, f"เกิดข้อผิดพลาดในการสร้างเวอร์ชัน: {e}")
            return redirect('manager:survey_list')
        else:
            messages.error(request, "เกิดข้อผิดพลาด กรุณาตรวจสอบข้อมูล")

    return redirect('manager:survey_list')

# (แถม) ฟังก์ชันลบ (ถ้ายังไม่มี)
@login_required
def survey_delete_view(request, pk):
    survey = get_object_or_404(Survey, pk=pk, service_point__in=request.user.managed_points.all())
    
    if request.method == 'POST':
        survey.delete()
        return redirect('manager:survey_list')

    return redirect('manager:survey_list')


@login_required
def question_list_view(request, survey_id):
    survey = get_object_or_404(Survey, pk=survey_id, service_point__in=request.user.managed_points.all())
    questions = survey.questions.all().order_by('order')
    show_modal = False

    if request.method == 'POST':
        form = ManagerQuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.survey = survey
            question.save()
            return redirect('manager:question_list', survey_id=survey.id)
        else:
            show_modal = True
            messages.error(request, "กรุณาตรวจสอบข้อมูล")
    else:
        # Auto Run Order: หาเลขลำดับถัดไป
        last_order = questions.count() + 1
        form = ManagerQuestionForm(initial={'order': last_order})

    context = {
        'survey': survey,
        'questions': questions,
        'form': form,
        'show_modal': show_modal,
        'page_title': f'จัดการคำถาม: {survey.title_th}'
    }

    return render(request, 'manager/question_list.html', context)

class QuestionUpdateView(LoginRequiredMixin, UpdateView):
    model = Question
    form_class = ManagerQuestionForm
    template_name = 'manager/question_form.html' 
    def get_queryset(self):
        return Question.objects.filter(
            survey__service_point__in=self.request.user.managed_points.all()
        )

    def get_success_url(self):
        messages.success(self.request, "แก้ไขคำถามเรียบร้อยแล้ว")
        return reverse('manager:question_list', args=[self.object.survey.id])

class QuestionDeleteView(LoginRequiredMixin, DeleteView):
    model = Question
    def get_queryset(self):
        # 🔒 Security Check
        return Question.objects.filter(
            survey__service_point__in=self.request.user.managed_points.all()
        )

    def get_success_url(self):
        messages.success(self.request, "ลบคำถามเรียบร้อยแล้ว")
        return reverse('manager:question_list', args=[self.object.survey.id])
 
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)

@login_required
def manager_assessment_results_view(request):
    user = request.user
    
    # 1. Base Security: ดึงเฉพาะจุดที่ User นี้ดูแล
    manager_points = user.managed_points.all()

    # 2. รับค่า Filter จาก URL
    group_id = request.GET.get('group_id')
    point_id = request.GET.get('point_id')
    score_filter = request.GET.get('score')
    
    # --- [ส่วนสำคัญ] จัดการวันที่ให้เหมือน Admin เป๊ะๆ ---
    end_date_str = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    start_date_str = request.GET.get('start_date', (timezone.now() - timedelta(days=6)).strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        # Fallback กรณีวันที่เพี้ยน
        start_date = (timezone.now() - timedelta(days=6)).date()
        end_date = timezone.now().date()
        # อัปเดต string ให้ตรงกับ date ที่ใช้จริง
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

    # บวก 1 วันเพื่อให้ครอบคลุมถึงเวลา 23:59:59 ของวันสิ้นสุด
    end_date_for_query = end_date + timedelta(days=1)
    # ----------------------------------------------------

    # 3. สร้าง Queryset หลัก (กรองตามจุดที่ดูแล + วันที่)
    responses = Response.objects.filter(
        service_point__in=manager_points,  # <--- กรองเฉพาะจุดของ Manager
        submitted_at__gte=start_date,
        submitted_at__lt=end_date_for_query
    ).annotate(
        avg_score=Avg('answers__answer_rating')
    ).select_related('service_point', 'service_point__group').order_by('-submitted_at')

    # 4. Apply Filters (กรองเพิ่มเติม)
    if group_id and group_id.isdigit():
        responses = responses.filter(service_point__group_id=int(group_id))
    
    # เช็คว่า point_id ที่ส่งมา เป็นจุดที่เขาดูแลจริงไหม (Security)
    if point_id and point_id.isdigit():
        if manager_points.filter(id=int(point_id)).exists():
            responses = responses.filter(service_point_id=int(point_id))

    if score_filter:
        try:
            min_score, max_score = map(int, score_filter.split('-'))
            if max_score == 5:
                responses = responses.filter(avg_score__gte=min_score, avg_score__lte=max_score)
            else:
                responses = responses.filter(avg_score__gte=min_score, avg_score__lt=max_score)
        except ValueError:
            pass

    # 5. Stats & Pagination
    total_assessments = responses.count()
    
    suggestion_queryset = ResponseAnswer.objects.filter(
        response__in=responses,
        question__question_type='TEXTAREA'
    ).exclude(answer_text='')
    
    total_suggestions = suggestion_queryset.count()
    recent_suggestions = suggestion_queryset.select_related('response', 'response__service_point').order_by('-id')[:10]

    paginator = Paginator(responses, 10) 
    page_obj = paginator.get_page(request.GET.get('page'))

    # 6. Prepare Choices (สำหรับ Dropdown)
    # แสดงเฉพาะกลุ่มภารกิจที่มีจุดบริการของ Manager อยู่
    group_ids = manager_points.values_list('group_id', flat=True).distinct()
    groups = ServiceGroup.objects.filter(id__in=group_ids).order_by('name')
    
    # สร้าง Map สำหรับ Dependent Dropdown
    point_map_json = json.dumps(_get_manager_point_map(manager_points))

    context = {
        'page_title': 'ผลการประเมิน (Manager)',
        'total_assessments': total_assessments,
        'total_suggestions': total_suggestions,
        'page_obj': page_obj,
        'recent_suggestions': recent_suggestions,
        'groups': groups,
        'points': manager_points, # ส่งจุดที่ดูแลไป
        'point_map_json': point_map_json,
        
        # ส่งค่ากลับไปที่ Form เพื่อให้หน้าเว็บจำค่าล่าสุด
        'selected_group': int(group_id) if group_id and group_id.isdigit() else '',
        'selected_point': int(point_id) if point_id and point_id.isdigit() else '',
        'selected_score': score_filter,
        'start_date': start_date_str, 
        'end_date': end_date_str,
    }
    return render(request, 'manager/assessment_results.html', context)

# ฟังก์ชันช่วยสร้าง Map (ต้องมีในไฟล์เดียวกัน หรือ import มา)
def _get_manager_point_map(manager_points_queryset):
    """สร้างแผนที่ (JSON) ของ {Group: [Points]} สำหรับ Manager"""
    point_map = {}
    # หา Group ที่เกี่ยวข้องกับ point ชุดนี้
    groups = ServiceGroup.objects.filter(
        service_points__in=manager_points_queryset
    ).distinct().prefetch_related('service_points')
    
    for group in groups:
        # ในแต่ละ Group, เอาเฉพาะ Point ที่ Manager ดูแล
        points_in_group = manager_points_queryset.filter(group=group).order_by('name')
        
        if points_in_group.exists():
            point_map[group.id] = [
                {'id': point.id, 'name': point.name}
                for point in points_in_group
            ]
    return point_map
@login_required
def suggestion_list_view(request): # เปลี่ยนชื่อฟังก์ชัน
    user = request.user
    manager_points = user.managed_points.all()
    
    # 1. รับค่า Filter (เหมือนหน้า Assessment)
    group_id = request.GET.get('group_id')
    point_id = request.GET.get('point_id')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    search_query = request.GET.get('q', '')

    suggestions = ResponseAnswer.objects.filter(
        response__service_point__in=manager_points, # 👈 การกรองหลักสำหรับ Manager
        question__question_type='TEXTAREA'
    ).exclude(answer_text='').select_related(
        'response', 
        'response__service_point', 
        'response__service_point__group'
    ).order_by('-response__submitted_at')
    
    # 3. Apply Filters (เหมือนเดิม)
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            suggestions = suggestions.filter(response__submitted_at__date__range=[start_date, end_date])
        except ValueError:
            pass

    if group_id:
        suggestions = suggestions.filter(response__service_point__group_id=group_id)
    if point_id:
        suggestions = suggestions.filter(response__service_point_id=point_id)
    
    if search_query:
        suggestions = suggestions.filter(answer_text__icontains=search_query)

    paginator = Paginator(suggestions, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    group_ids = manager_points.values_list('group_id', flat=True).distinct()
    groups = ServiceGroup.objects.filter(id__in=group_ids)
    
    points = manager_points # ใช้ managed_points ที่กรองไว้แล้ว
    if group_id: points = points.filter(group_id=group_id)
    
    context = {
        'page_title': 'รายการข้อเสนอแนะทั้งหมด',
        'page_obj': page_obj,
        'groups': groups,
        'points': points,
        'selected_group': int(group_id) if group_id else '',
        'selected_point': int(point_id) if point_id else '',
        'start_date': start_date_str if start_date_str else '',
        'end_date': end_date_str if end_date_str else '',
        'search_query': search_query,
    }
    return render(request, 'manager/suggestion_list.html', context)


# --- Helper สำหรับ Manager (กรองเฉพาะจุดที่ดูแล) ---
def _get_manager_filtered_responses(request):
    user = request.user
    
    # 1. Key Logic: ดึงเฉพาะจุดบริการที่ User นี้ดูแล
    managed_points = user.managed_points.all()
    
    # 2. Date Filter
    default_end = timezone.now().date()
    default_start = (timezone.now() - timedelta(days=30)).date()
    
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else default_end
    except (ValueError, TypeError):
        start_date = default_start
        end_date = default_end

    end_date_query = end_date + timedelta(days=1)

    # 3. Query Responses (กรองด้วย managed_points)
    responses = Response.objects.filter(
        service_point__in=managed_points,  
        submitted_at__gte=start_date,
        submitted_at__lt=end_date_query
    )

    # 4. Filter เพิ่มเติมจากหน้าเว็บ (ถ้า Manager เลือกเจาะจงจุดใดจุดหนึ่ง)
    point_id = request.GET.get('point_id')
    if point_id and point_id.isdigit():
        # ต้องเช็คด้วยว่า point_id ที่ส่งมา อยู่ในสิทธิ์ที่เขาดูแลไหม (กันเหนียว)
        if managed_points.filter(id=int(point_id)).exists():
            responses = responses.filter(service_point_id=int(point_id))
        
    return responses

# ==========================================
# Manager Export: Dashboard Summary
# ==========================================
def export_manager_dashboard_summary(request):
    """
    ฟังก์ชัน Export รายงานสรุปสำหรับ Manager (ฉบับสมบูรณ์ แก้ไขข้อมูลหาย)
    """
    
    # 1. ดึงข้อมูล (Logic การกรอง)
    responses = _get_manager_filtered_responses(request)
    managed_points = request.user.managed_points.all()
    
    # จัดการวันที่สำหรับ Header
    default_end = timezone.now().date()
    default_start = (timezone.now() - timedelta(days=30)).date()
    
    req_start = request.GET.get('start_date')
    req_end = request.GET.get('end_date')

    show_start_date = req_start if req_start else default_start.strftime('%Y-%m-%d')
    show_end_date = req_end if req_end else default_end.strftime('%Y-%m-%d')
    
    # ---------------------------------------------------------
    # 2. คำนวณตัวเลข (Calculation)
    # ---------------------------------------------------------
    
    # 2.1 KPI
    total_responses = responses.count()
    total_my_points = managed_points.count()
    
    active_questions = Question.objects.filter(
        survey__service_point__in=managed_points,
        survey__status='ACTIVE'
    ).count()

    # 2.2 Top Service Points (เฉพาะในกลุ่มที่ดูแล)
    sp_stats = responses.values('service_point__name')\
        .annotate(total=Count('id'))\
        .order_by('-total')

    # 2.3 Weekly Stats (Pure Python)
    raw_datetimes = responses.values_list('submitted_at', flat=True)
    weekly_stats = {}
    for dt in raw_datetimes:
        if dt is None: continue
        local_date = timezone.localtime(dt).date()
        monday = local_date - timedelta(days=local_date.weekday())
        if monday not in weekly_stats: weekly_stats[monday] = 0
        weekly_stats[monday] += 1
    
    weekly_stats_list = [{'week': k, 'total': v} for k, v in weekly_stats.items()]
    weekly_stats_list.sort(key=lambda x: x['week'])

    # ---------------------------------------------------------
    # 3. เขียนลง Excel (Writing to Excel)
    # ---------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manager Summary"
    
    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    
    # --- HEADER ---
    ws.append(["รายงานสรุปผลการประเมิน (สำหรับผู้ดูแล)"])
    ws['A1'].font = header_font
    
    ws.append([f"ผู้ดูแล: {request.user.get_full_name() or request.user.username}"])
    ws.append([f"ช่วงเวลาข้อมูล: {show_start_date} ถึง {show_end_date}"])
    ws.append([]) 

    # --- SECTION 1: KPI (เพิ่มส่วนนี้ที่หายไปกลับมา) ---
    ws.append(["1. ภาพรวม (KPIs)"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    
    ws.append(["หัวข้อ", "จำนวน"])
    ws.append(["จำนวนการตอบทั้งหมด", total_responses])
    ws.append(["จำนวนจุดบริการที่ดูแล", total_my_points])
    ws.append(["จำนวนคำถาม (Active)", active_questions])
    ws.append([])

    # --- SECTION 2: STATS BY POINT (เพิ่มส่วนนี้ที่หายไปกลับมา) ---
    ws.append(["2. สถิติการประเมินแยกตามจุดบริการ"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    
    ws.append(["ชื่อจุดบริการ", "จำนวนครั้ง"])
    if not sp_stats:
        ws.append(["ไม่มีการประเมินในช่วงเวลานี้", "-"])
    else:
        for item in sp_stats:
            ws.append([item['service_point__name'], item['total']])
    ws.append([])

    # --- SECTION 3: WEEKLY TREND (เพิ่มส่วนนี้ที่หายไปกลับมา) ---
    ws.append(["3. แนวโน้มรายสัปดาห์"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    
    ws.append(["สัปดาห์ (เริ่มวันจันทร์)", "จำนวนการประเมิน"])
    if not weekly_stats_list:
        ws.append(["ไม่พบข้อมูล", "-"])
    else:
        for item in weekly_stats_list:
            week_str = item['week'].strftime('%Y-%m-%d')
            ws.append([week_str, item['total']])
    ws.append([])
    
    # --- SECTION 4: MANAGED POINTS ---
    ws.append(["4. จุดบริการที่ท่านรับผิดชอบ"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    for p in managed_points:
        ws.append([p.name])

    # จัดความกว้าง
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="manager_summary.xlsx"'
    wb.save(response)
    return response

# ==========================================
# Manager Export: Raw Data & Suggestion
# ==========================================
# ใช้ Logic เดียวกับ Superuser แต่เปลี่ยนไปเรียก _get_manager_filtered_responses

def export_manager_assessment_excel(request):
    # 1. ใช้ Helper ของ Manager เพื่อดึงข้อมูลที่กรองเฉพาะจุดที่ดูแล
    responses = _get_manager_filtered_responses(request)
    
    # 2. Query ข้อมูลคำตอบ
    queryset = ResponseAnswer.objects.filter(response__in=responses)\
        .select_related('response', 'response__service_point', 'question')\
        .order_by('response__submitted_at')

    # 3. สร้าง Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Data"
    
    headers = ['Response ID', 'Service Point', 'Submitted At', 'Role', 'Question', 'Answer (Rating/Text)']
    ws.append(headers)
    
    # จัดความกว้าง
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        col_letter = get_column_letter(col_num)
        
        if col_num == 2: ws.column_dimensions[col_letter].width = 30 # Service Point
        elif col_num == 3: ws.column_dimensions[col_letter].width = 25 # Date
        elif col_num == 5: ws.column_dimensions[col_letter].width = 50 # Question
        else: ws.column_dimensions[col_letter].width = 20

    for ans in queryset:
        local_time = timezone.localtime(ans.response.submitted_at).replace(tzinfo=None)
        
        # เลือกค่าที่จะแสดง
        val = ans.answer_rating if ans.answer_rating is not None else ans.answer_text
        
        ws.append([
            ans.response.id,
            ans.response.service_point.name,
            local_time,
            ans.response.user_role,
            getattr(ans.question, 'text_th', ''),
            val
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="manager_assessment_data.xlsx"'
    wb.save(response)
    return response

def export_manager_suggestion_excel(request):
    # 1. ใช้ Helper ของ Manager
    responses = _get_manager_filtered_responses(request)
    
    # 2. กรองเฉพาะข้อเสนอแนะ (TEXTAREA) และไม่เอาค่าว่าง
    queryset = ResponseAnswer.objects.filter(
        response__in=responses,
        question__question_type='TEXTAREA'
    ).exclude(answer_text='')\
    .select_related('response', 'response__service_point')\
    .order_by('response__submitted_at')

    # 3. สร้าง Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suggestions"
    
    headers = ['Date/Time', 'Service Point', 'Group', 'User Role', 'Suggestion']
    ws.append(headers)
    
    # จัดความกว้าง
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        col_letter = get_column_letter(col_num)

        if col_num == 1: ws.column_dimensions[col_letter].width = 25 # Date
        elif col_num == 2: ws.column_dimensions[col_letter].width = 30 # Point
        elif col_num == 5: ws.column_dimensions[col_letter].width = 60 # Suggestion
        else: ws.column_dimensions[col_letter].width = 20

    for ans in queryset:
        local_time = timezone.localtime(ans.response.submitted_at).replace(tzinfo=None)
        
        ws.append([
            local_time,
            ans.response.service_point.name,
            ans.response.service_point.group.name,
            ans.response.user_role,
            ans.answer_text
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="manager_suggestion_list.xlsx"'
    wb.save(response)
    return response

def export_manager_assessment_csv(request):
    responses = _get_manager_filtered_responses(request)
    queryset = ResponseAnswer.objects.filter(response__in=responses)\
        .select_related('response', 'response__service_point', 'question')\
        .order_by('response__submitted_at')

    response = HttpResponse(content_type='text/csv', headers={'Content-Disposition': 'attachment; filename="manager_assessment_data.csv"'})
    response.write('\ufeff') 
    writer = csv.writer(response)
    writer.writerow(['Response ID', 'Service Point', 'Submitted At', 'Role', 'Question', 'Answer (Rating/Text)'])
    
    for ans in queryset:
        val = ans.answer_rating if ans.answer_rating is not None else ans.answer_text
        writer.writerow([
            ans.response.id,
            ans.response.service_point.name,
            ans.response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            ans.response.user_role,
            getattr(ans.question, 'text_th', ''),
            val
        ])
    return response

def export_manager_suggestion_csv(request):
    responses = _get_manager_filtered_responses(request)
    queryset = ResponseAnswer.objects.filter(
        response__in=responses, question__question_type='TEXTAREA'
    ).exclude(answer_text='').select_related('response', 'response__service_point').order_by('response__submitted_at')

    response = HttpResponse(content_type='text/csv', headers={'Content-Disposition': 'attachment; filename="manager_suggestion_list.csv"'})
    response.write('\ufeff') 
    writer = csv.writer(response)
    writer.writerow(['Date/Time', 'Service Point', 'Group', 'User Role', 'Suggestion'])
    
    for ans in queryset:
        writer.writerow([
            ans.response.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            ans.response.service_point.name,
            ans.response.service_point.group.name,
            ans.response.user_role,
            ans.answer_text
        ])
    return response



from django.views.decorators.http import require_POST
from django.http import JsonResponse
@login_required
@require_POST
def clear_all_notifications(request):
    Notification.objects.filter(recipient=request.user).delete()
    return JsonResponse({'status': 'success'})